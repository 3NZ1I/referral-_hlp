
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import jwt
import logging
import traceback
from datetime import datetime, timedelta, timezone
from sqlalchemy.orm import Session
from sqlalchemy.exc import OperationalError
from backend.models import User, Case, Comment, ImportJob, ImportRow
from datetime import datetime
from .schemas import (
    CaseCreate,
    CaseUpdate,
    CaseRead,
    UserCreate,
    UserRead,
    CommentCreate,
    CommentRead,
)
from sqlalchemy import create_engine, or_
from sqlalchemy.orm import sessionmaker
import os
import openpyxl
import bcrypt
import json

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+psycopg2://user:password@localhost/referral_db")
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+psycopg2://user:password@localhost/referral_db")
try:
    # Provide an explicit, more helpful error if the DB driver is missing
    if DATABASE_URL.startswith('postgresql'):
        try:
            import psycopg2  # noqa: F401
        except ModuleNotFoundError:
            logging.error('psycopg2 driver not found in Python environment. Install via: pip install psycopg2-binary')
            raise RuntimeError('psycopg2 driver not found; please install psycopg2-binary or set DATABASE_URL to use sqlite for local testing')
    engine = create_engine(DATABASE_URL)
except Exception as e:
    logging.exception('Failed to create database engine: %s', e)
    # Re-raise so startup fails explicitly with a clearer message
    raise
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

app = FastAPI()

# Configure CORS
_allowed_origins = os.getenv('CORS_ORIGINS', '*')
if _allowed_origins and _allowed_origins != '*':
    try:
        # split comma-separated origins
        _origins_list = [o.strip() for o in _allowed_origins.split(',') if o.strip()]
    except Exception:
        _origins_list = [_allowed_origins]
else:
    _origins_list = ['*']

# Track whether we allow credentials (this matters for Access-Control-Allow-Origin header)
_cors_allow_credentials = True

app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def ensure_cors_headers(request: Request, call_next):
    """Ensure Access-Control headers are set and if allow_credentials=True we echo the request origin
    to avoid sending wildcard '*' which is invalid when credentials are included by the browser.
    """
    response = await call_next(request)
    origin = request.headers.get('origin')
    try:
        if origin:
            if ('*' in _origins_list) or (origin in _origins_list):
                # If credentials are allowed, echo the origin explicitly and include credentials header
                if _cors_allow_credentials:
                    response.headers['Access-Control-Allow-Origin'] = origin
                    response.headers['Access-Control-Allow-Credentials'] = 'true'
                else:
                    # Non-credentialed responses may use a wildcard if allowed
                    if '*' in _origins_list:
                        response.headers['Access-Control-Allow-Origin'] = '*'
                    else:
                        response.headers['Access-Control-Allow-Origin'] = origin
            else:
                # If origin not in allowed list, don't add header
                pass
        else:
            # no origin provided; ensure a fallback if missing
            if 'Access-Control-Allow-Origin' not in response.headers:
                response.headers['Access-Control-Allow-Origin'] = '*'
    except Exception:
        # Be defensive if something goes wrong while setting headers
        if 'Access-Control-Allow-Origin' not in response.headers:
            response.headers['Access-Control-Allow-Origin'] = '*'
    return response
MAINTENANCE_SCHEDULES = []
_maintenance_seq = 1

def _normalize_roles(user):
    """Return a normalized list of role strings for the given user token.
    The token may supply `role` as a string, a list `roles`, or not at all.
    """
    if not isinstance(user, dict):
        return []
    raw = user.get('roles') or user.get('role') or user.get('roles')
    if not raw:
        return []
    if isinstance(raw, str):
        return [raw.strip().lower()]
    if isinstance(raw, (list, tuple)):
        return [str(r).strip().lower() for r in raw if r is not None]
    # otherwise, attempt to coerce to string
    return [str(raw).strip().lower()]


def is_admin_user(user):
    roles = _normalize_roles(user)
    return 'admin' in roles


def _ensure_submission_time_in_raw(raw: dict) -> dict:
    """If raw contains a `body` object with submission timestamp fields, copy them to top-level `raw`.
    This helps the frontend compute Age consistently when a wrapper payload was received.
    """
    if not isinstance(raw, dict):
        return raw
    try:
        body = raw.get('body')
        if isinstance(body, dict):
            for field in ('_submission_time', 'submissiontime', 'submission_time'):
                if body.get(field) is not None and raw.get(field) is None:
                    raw[field] = body.get(field)
                    try:
                        del body[field]
                    except Exception:
                        pass
    except Exception:
        pass

    # Normalize known submission timestamp fields (int epoch or common string formats) into an ISO 8601 string
    try:
        for field in ('_submission_time', 'submissiontime', 'submission_time'):
            val = raw.get(field)
            if val is None:
                continue
            # If already an int/float timestamp, convert to ISO UTC
            if isinstance(val, (int, float)):
                # If milliseconds (large number), normalize to seconds
                ts = int(val)
                if ts > 1e12:  # milliseconds
                    ts = ts // 1000
                raw[field] = datetime.utcfromtimestamp(ts).isoformat() + 'Z'
                continue
            if isinstance(val, str):
                s = val.strip()
                # purely numeric string: interpret as unix seconds or ms
                if s.isdigit():
                    ts = int(s)
                    if ts > 1e12:  # ms
                        ts = ts // 1000
                    raw[field] = datetime.utcfromtimestamp(ts).isoformat() + 'Z'
                    continue
                # Normal ISO-like strings with space instead of T: replace and append Z if missing
                try:
                    if ' ' in s and 'T' not in s and 'Z' not in s:
                        s2 = s.replace(' ', 'T') + 'Z'
                    else:
                        s2 = s
                    # Try parsing
                    dt = datetime.fromisoformat(s2.replace('Z', '+00:00'))
                    # Re-serialize as UTC ISO with Z
                    raw[field] = dt.astimezone(timezone.utc).isoformat().replace('+00:00', 'Z')
                except Exception:
                    # Last resort: leave as-is since frontend may accept it
                    pass
    except Exception:
        # Be resilient: do nothing on parse errors
        pass
    return raw

def has_role(user, role_name):
    if not role_name:
        return False
    roles = _normalize_roles(user)
    return role_name.strip().lower() in roles


# Lightweight health endpoint for readiness checks
@app.get("/health")
def health():
    return {"status": "ok"}


 
@app.exception_handler(Exception)
def general_exception_handler(request, exc):
    # Log full stack trace locally for diagnostics, avoid exposing details in response
    logging.exception("Unhandled exception in API request: %s", exc)
    # Ensure CORS header presence on error responses so browser can receive the error
    # If request includes an Origin, echo it to avoid wildcard when credentials are used
    origin = None
    try:
        origin = request.headers.get('origin')
    except Exception:
        origin = None
    headers = {}
    if origin:
        headers["Access-Control-Allow-Origin"] = origin
        headers["Access-Control-Allow-Credentials"] = "true"
    else:
        headers["Access-Control-Allow-Origin"] = "*"
        headers["Access-Control-Allow-Credentials"] = "true"
    return JSONResponse({"detail": "Internal server error"}, status_code=500, headers=headers)


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    try:
        body = await request.body()
    except Exception:
        body = None
    logging.warning('Request validation failed: url=%s errors=%s body=%s', request.url, exc.errors(), body)
    # Mirror the default 422 response but ensure we log body for diagnostics
    return JSONResponse(status_code=422, content={"detail": exc.errors()})

security = HTTPBearer()
optional_security = HTTPBearer(auto_error=False)
JWT_SECRET = os.getenv("SECRET_KEY", os.getenv("JWT_SECRET", "dev-secret"))
JWT_EXP_MINUTES = int(os.getenv("JWT_EXP_MINUTES", "120"))

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))


def validate_password_strength(password: str) -> bool:
    # Enforce minimum strength: at least 8 characters, at least one lowercase, one uppercase, and one digit/special
    if not password or len(password) < 8:
        return False
    has_lower = any(c.islower() for c in password)
    has_upper = any(c.isupper() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_symbol = any(not c.isalnum() for c in password)
    return has_lower and has_upper and (has_digit or has_symbol)


def create_token(payload: dict, exp_minutes: int | None = None):
    """Create a JWT token with an optional custom expiry (in minutes).
    If exp_minutes is None, defaults to JWT_EXP_MINUTES from env.
    """
    to_encode = payload.copy()
    minutes = JWT_EXP_MINUTES if exp_minutes is None else int(exp_minutes)
    to_encode.update({"exp": datetime.utcnow() + timedelta(minutes=minutes)})
    return jwt.encode(to_encode, JWT_SECRET, algorithm="HS256")


def require_auth(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        decoded = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        return decoded
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or expired token")


def optional_auth(credentials: HTTPAuthorizationCredentials = Depends(optional_security)):
    if not credentials:
        return None
    try:
        token = credentials.credentials
        decoded = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        return decoded
    except Exception:
        return None

 

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# CASES CRUD
@app.get("/cases", response_model=list[CaseRead])
def get_cases(db: Session = Depends(get_db), user=Depends(optional_auth)):
    try:
        cases = db.query(Case).all()
    except OperationalError as e:
        logging.exception('Database connection failed while fetching cases: %s', e)
        raise HTTPException(status_code=503, detail='Database unavailable')
    # Hide sensitive fields in the raw payload for non-admin users
    sensitive_keys = set(['id_card_nu', 'family_card_nu', 'passport_nu_001', 'passaport_nu_001'])
    if not is_admin_user(user):
        for c in cases:
            if isinstance(c.raw, dict):
                sanitized = dict(c.raw)
                for k in list(sanitized.keys()):
                    if k in sensitive_keys:
                        sanitized.pop(k, None)
                c.raw = sanitized
            # Normalize empty emails on assigned_to relation
            if c.assigned_to and isinstance(c.assigned_to.email, str) and c.assigned_to.email.strip() == '':
                c.assigned_to.email = None
    return cases

@app.get("/cases/{case_id}", response_model=CaseRead)
def get_case(case_id: int, db: Session = Depends(get_db), user=Depends(optional_auth)):
    try:
        case = db.get(Case, case_id)
    except OperationalError as e:
        logging.exception('Database connection failed while fetching case %s: %s', case_id, e)
        raise HTTPException(status_code=503, detail='Database unavailable')
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    # Sanitize raw fields for non-admins
    sensitive_keys = set(['id_card_nu', 'family_card_nu', 'passport_nu_001', 'passaport_nu_001'])
    if not is_admin_user(user) and isinstance(case.raw, dict):
        sanitized = dict(case.raw)
        for k in list(sanitized.keys()):
            if k in sensitive_keys:
                sanitized.pop(k, None)
        case.raw = sanitized
    if case.assigned_to and isinstance(case.assigned_to.email, str) and case.assigned_to.email.strip() == '':
        case.assigned_to.email = None
    return case

@app.post("/cases", response_model=CaseRead, status_code=status.HTTP_201_CREATED)
def create_case(case: CaseCreate, db: Session = Depends(get_db), user=Depends(require_auth)):
    # Backed-up original code:
    # new_case = Case(**case.dict())
    # Sanitize and normalize incoming payload before creating the case
    payload = case.dict()
    raw = payload.get('raw')
    try:
        # If raw was received as a JSON string, parse it
        if isinstance(raw, str) and raw.strip():
            try:
                raw = json.loads(raw)
            except Exception:
                # keep the raw string in a 'body' key if not JSON
                raw = {'body': raw}
        # If raw contains `body` that is stringified JSON, parse it
        if isinstance(raw, dict) and isinstance(raw.get('body'), str):
            try:
                raw['body'] = json.loads(raw['body'])
            except Exception:
                pass
    except Exception:
        raw = payload.get('raw')
    payload['raw'] = raw
    # Ensure submission timestamps are available at top-level raw for frontend Age computation
    raw = _ensure_submission_time_in_raw(payload['raw'])
    payload['raw'] = raw
    try:
        if isinstance(raw, dict) and 'headers' in raw and isinstance(raw.get('body'), dict) and len(raw) <= 2:
            flattened = dict(raw.get('body'))
            flattened['headers'] = raw.get('headers')
            # move any kobo_* fields to flattened
            if raw.get('kobo_case_id'):
                flattened['kobo_case_id'] = raw.get('kobo_case_id')
            raw = flattened
            payload['raw'] = raw
    except Exception:
        pass

    # Compute title if missing or unhelpful (e.g. 'Kobo Submission')
    title = payload.get('title') or ''
    def _first_nonempty(*vals):
        for v in vals:
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == '':
                continue
            return v
        return None
    # Try to find case number in various known locations
    case_number = None
    if isinstance(raw, dict):
        case_number = _first_nonempty(
            raw.get('case_number'),
            raw.get('caseNumber'),
            (raw.get('body') or {}).get('case_number') if isinstance(raw.get('body'), dict) else None,
            payload.get('caseNumber')
        )
    # If title is blank or default value, prefer case_number or beneficiary name
    if not title or title.strip() == '' or title.strip().lower() == 'kobo submission':
        if case_number:
            title = str(case_number)
        else:
            beneficiary = _first_nonempty((raw or {}).get('beneficiary_name'), (raw or {}).get('name'))
            title = str(beneficiary) if beneficiary else (payload.get('title') or f'Kobo {raw.get("_uuid") if raw and isinstance(raw, dict) else "submission"}')
    payload['title'] = title
    new_case = Case(**payload)
    db.add(new_case)
    db.commit()
    db.refresh(new_case)
    logging.info('Created case via API: id=%s title=%s', new_case.id, new_case.title)
    return new_case

@app.put("/cases/{case_id}", response_model=CaseRead)
def update_case(case_id: int, case: CaseUpdate, db: Session = Depends(get_db), user=Depends(require_auth)):
    db_case = db.get(Case, case_id)
    if not db_case:
        raise HTTPException(status_code=404, detail="Case not found")
    # enforce resolve comment when changing status to resolved states
    payload = case.dict(exclude_unset=True)
    # Sanitize raw if present in payload
    raw_payload = payload.get('raw')
    if raw_payload is not None:
        try:
            if isinstance(raw_payload, str) and raw_payload.strip():
                try:
                    raw_payload = json.loads(raw_payload)
                except Exception:
                    raw_payload = {'body': raw_payload}
            if isinstance(raw_payload, dict) and isinstance(raw_payload.get('body'), str):
                try:
                    raw_payload['body'] = json.loads(raw_payload['body'])
                except Exception:
                    pass
        except Exception:
            pass
        payload['raw'] = raw_payload
    # If the title provided was a generic Kobo placeholder, try to set a useful title
    if 'title' in payload:
        t = payload.get('title') or ''
        if t.strip().lower() == 'kobo submission' or t.strip() == '':
            # find a case number in raw or body
            raw = payload.get('raw')
            def _first_nonempty(*vals):
                for v in vals:
                    if v is None:
                        continue
                    if isinstance(v, str) and v.strip() == '':
                        continue
                    return v
                return None
            if isinstance(raw, dict):
                case_number = _first_nonempty(raw.get('case_number'), raw.get('caseNumber'), (raw.get('body') or {}).get('case_number') if isinstance(raw.get('body'), dict) else None)
                if case_number:
                    payload['title'] = str(case_number)
    # Flatten wrapper-style raw if needed
    raw_upd = payload.get('raw')
    try:
        if isinstance(raw_upd, dict) and 'headers' in raw_upd and isinstance(raw_upd.get('body'), dict) and len(raw_upd) <= 2:
            flattened = dict(raw_upd.get('body'))
            flattened['headers'] = raw_upd.get('headers')
            if raw_upd.get('kobo_case_id'):
                flattened['kobo_case_id'] = raw_upd.get('kobo_case_id')
            payload['raw'] = flattened
            # Ensure top-level submission time fields are set if present in nested body
            payload['raw'] = _ensure_submission_time_in_raw(payload['raw'])
    except Exception:
        pass
    new_status = payload.get('status')
    resolved_states = ['Completed', 'Closed']
    if new_status and new_status in resolved_states and new_status != (db_case.status or ''):
        resolve_comment = payload.get('resolve_comment')
        if not resolve_comment or not resolve_comment.strip():
            raise HTTPException(status_code=400, detail='Resolve comment is required when changing status to a resolved state')
        # create a comment for the resolve action
        try:
            user_id = None
            if isinstance(user, dict):
                user_id = user.get('user_id') or user.get('sub')
            cmt = Comment(case_id=case_id, user_id=user_id, content=resolve_comment)
            db.add(cmt)
            db.commit()
        except Exception as e:
            logging.exception('Failed to persist resolve comment: %s', e)
            db.rollback()
    # apply updates to case
    for key, value in payload.items():
        if key == 'resolve_comment':
            continue
        setattr(db_case, key, value)
    # Update timestamps
    try:
        db_case.updated_at = datetime.utcnow()
        if new_status and new_status in resolved_states:
            db_case.completed_at = datetime.utcnow()
    except Exception:
        logging.exception('Failed to set updated_at or completed_at on case')
    db.commit()
    db.refresh(db_case)
    return db_case

@app.delete("/cases/{case_id}")
def delete_case(case_id: int, db: Session = Depends(get_db), user=Depends(require_auth)):
    db_case = db.get(Case, case_id)
    if not db_case:
        raise HTTPException(status_code=404, detail="Case not found")
    print(f'DEBUG: Attempt delete case {case_id} by user {user}')
    logging.info('Attempt delete case %s by user %s', case_id, user)
    roles = _normalize_roles(user)
    logging.info('is_admin_user=%s, roles=%s', is_admin_user(user), roles)
    # permission check: only admin or internal
    if not (is_admin_user(user) or has_role(user, 'internal')):
        raise HTTPException(status_code=403, detail='Insufficient permissions to delete case')
    try:
        # Delete comments related to this case to avoid FK constraint issues
        deleted_comments = db.query(Comment).filter(Comment.case_id == case_id).delete(synchronize_session=False)
        # Nullify import rows referencing this case
        updated_import_rows = db.query(ImportRow).filter(ImportRow.case_id == case_id).update({ImportRow.case_id: None}, synchronize_session=False)
        db.add(db_case)
        db.delete(db_case)
        db.commit()
        return {"detail": "Case deleted", 'deleted_comments': deleted_comments, 'updated_import_rows': updated_import_rows}
    except Exception as e:
        db.rollback()
        logging.exception('Failed to delete case %s: %s', case_id, e)
        raise HTTPException(status_code=500, detail='Internal server error')

# USERS CRUD
@app.get("/users", response_model=list[UserRead])
def get_users(db: Session = Depends(get_db)):
    try:
        # explicit DB connectivity check: handle OperationalError to return 503
        users = db.query(User).all()
        # Normalize empty email strings to None to avoid Pydantic EmailStr validation errors
        for u in users:
            if isinstance(u.email, str) and u.email.strip() == '':
                u.email = None
        logging.info('GET /users returned %s users', len(users))
        return users
    except OperationalError as e:
        logging.exception('Database connection failed while fetching users: %s', e)
        raise HTTPException(status_code=503, detail='Database unavailable')
    except Exception as e:
        logging.exception('Failed to fetch users: %s', e)
        raise HTTPException(status_code=500, detail='Internal server error (could not fetch users)')

@app.post("/users", response_model=UserRead, status_code=status.HTTP_201_CREATED)
def create_user(user: UserCreate, db: Session = Depends(get_db), auth=Depends(require_auth)):
    # Hash password if provided
    password = getattr(user, 'password', None)
    if password and not validate_password_strength(password):
        raise HTTPException(status_code=400, detail='Password does not meet strength requirements (min 8 chars; mixed case; digits or symbols)')
    user_data = user.dict(exclude={"password"})
    # Normalize empty string email to None to avoid Pydantic EmailStr validation issues
    if 'email' in user_data and isinstance(user_data['email'], str) and user_data['email'].strip() == '':
        user_data['email'] = None
    if password:
        user_data["password_hash"] = hash_password(password)
    # Accept must_change_password if provided by admin
    if 'must_change_password' in user.dict():
        user_data['must_change_password'] = user.dict().get('must_change_password')
    new_user = User(**user_data)
    try:
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        logging.info('Created user id=%s username=%s', new_user.id, new_user.username)
        return new_user
    except Exception as e:
        logging.exception('Failed to create user: %s', e)
        db.rollback()
        raise HTTPException(status_code=500, detail='Internal server error while creating user')


@app.delete('/users/{user_id}', status_code=status.HTTP_204_NO_CONTENT)
def delete_user(user_id: int, db: Session = Depends(get_db), auth=Depends(require_auth)):
    db_user = db.get(User, user_id)
    if not db_user:
        raise HTTPException(status_code=404, detail='User not found')
    db.delete(db_user)
    db.commit()
    return JSONResponse(status_code=status.HTTP_204_NO_CONTENT)


@app.put('/users/{user_id}', response_model=UserRead)
def update_user(user_id: int, payload: dict, db: Session = Depends(get_db), auth=Depends(require_auth)):
    db_user = db.get(User, user_id)
    if not db_user:
        raise HTTPException(status_code=404, detail='User not found')
    # Only allow admin or the user itself to update
    if not is_admin_user(auth) and (not isinstance(auth, dict) or auth.get('user_id') != user_id):
        raise HTTPException(status_code=403, detail='Insufficient permissions')

    # Password change needs validation
    password = payload.get('password')
    if password:
        if not validate_password_strength(password):
            raise HTTPException(status_code=400, detail='Password does not meet strength requirements (min 8 chars; mixed case; digits or symbols)')
        db_user.password_hash = hash_password(password)

    # Update other allowed fields: email, name, role, ability, must_change_password
    for k in ['email', 'name', 'role', 'ability', 'must_change_password']:
        if k in payload:
            val = payload.get(k)
            if k == 'email' and isinstance(val, str) and val.strip() == '':
                val = None
            setattr(db_user, k, val)
    db.commit()
    db.refresh(db_user)
    return db_user

# COMMENTS CRUD
@app.get("/cases/{case_id}/comments", response_model=list[CommentRead])
def get_comments(case_id: int, db: Session = Depends(get_db)):
    return db.query(Comment).filter(Comment.case_id == case_id).all()

@app.post("/cases/{case_id}/comments", response_model=CommentRead, status_code=status.HTTP_201_CREATED)
def add_comment(case_id: int, comment: CommentCreate, db: Session = Depends(get_db), user=Depends(require_auth)):
    # Use the authenticated user id for comment ownership when possible
    user_id = None
    if isinstance(user, dict):
        user_id = user.get('user_id') or user.get('sub')
    new_comment = Comment(case_id=case_id, user_id=user_id, **comment.dict())
    db.add(new_comment)
    # Update the case's updated_at timestamp when adding a new comment
    try:
        existing_case = db.get(Case, case_id)
        if existing_case:
            existing_case.updated_at = datetime.utcnow()
            db.add(existing_case)
    except Exception:
        logging.exception('Failed to set updated_at for case when adding comment')
    db.commit()
    db.refresh(new_comment)
    return new_comment

# ABILITIES (simple list)
@app.get("/abilities")
def get_abilities(db: Session = Depends(get_db)):
    abilities = db.query(User.ability).distinct().all()
    return [a[0] for a in abilities if a[0]]

# ASSIGN CASE
@app.post("/cases/{case_id}/assign", response_model=CaseRead)
def assign_case(case_id: int, payload: dict, db: Session = Depends(get_db), user=Depends(require_auth)):
    case = db.get(Case, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    user_name = payload.get("user")
    ability = payload.get("ability")
    assigned_user = None
    if user_name:
        assigned_user = db.query(User).filter(User.name == user_name).first()
        if not assigned_user:
            assigned_user = User(name=user_name, ability=ability)
            db.add(assigned_user)
            db.flush()
    elif ability:
        assigned_user = db.query(User).filter(User.ability == ability).first()
        if not assigned_user:
            assigned_user = User(name=f"auto-{ability}", ability=ability)
            db.add(assigned_user)
            db.flush()
    # Prefer to set assigned_to_id explicitly to ensure DB-level FK sync
    case.assigned_to = assigned_user
    case.assigned_to_id = assigned_user.id if assigned_user else None
    # Update timestamp and add an assignment comment for audit trail
    try:
        case.updated_at = datetime.utcnow()
        db.add(case)
        user_id = None
        if isinstance(user, dict):
            user_id = user.get('user_id') or user.get('sub')
        assign_text = f"Assigned to {assigned_user.name if assigned_user else 'Unassigned'}"
        comment = Comment(case_id=case_id, user_id=user_id, content=assign_text)
        db.add(comment)
    except Exception:
        logging.exception('Failed to set updated_at or persist assignment comment')
    db.commit()
    db.refresh(case)
    return case

# XLSX IMPORT (n8n/file upload compatible)
@app.post("/import")
def import_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(require_auth)):
    try:
        wb = openpyxl.load_workbook(file.file)
        # continue with import logic
    except HTTPException:
        # Re-raise FastAPI HTTPException as-is
        raise
    except Exception as e:
        logging.exception('Unhandled exception while parsing uploaded file: %s', e)
        raise HTTPException(status_code=400, detail=f"Invalid XLSX file: {e}")
    ws = wb.active
    try:
        header_row = ws[1]
        headers = [cell.value for cell in header_row]
        if not any(headers):
            raise HTTPException(status_code=400, detail='XLSX header row is empty - no column names were detected')
    except IndexError:
        # No rows in sheet
        raise HTTPException(status_code=400, detail='XLSX file contains no rows')
    except HTTPException:
        raise
    except Exception as e:
        logging.exception('Failed to parse header row in XLSX file: %s', e)
        raise HTTPException(status_code=400, detail=f'Unable to parse header row: {e}')
    imported = 0
    created_ids = []
    failed_rows = []
    # Resolve uploader id if authentication provided (ensure uploader_user is known before creating Job)
    uploader_user = None
    if isinstance(user, dict):
        uploader_id = user.get('user_id') or user.get('sub')
        if uploader_id:
            try:
                uploader_user = db.get(User, int(uploader_id))
            except Exception:
                uploader_user = None

    # create an import job record
    job = ImportJob(
        uploader_id=uploader_user.id if uploader_user else None,
        uploader_name=(uploader_user.name if uploader_user else None),
        filename=getattr(file, 'filename', None)
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    # (moved above) uploader_user already resolved so proceed

    def sanitize_value(v):
        # Ensure JSON serializable; convert datetime to ISO strings, bytes to str.
        try:
            json.dumps(v)
            return v
        except Exception:
            if isinstance(v, (datetime,)):
                return v.isoformat()
            return str(v)

    def sanitize_obj(obj):
        if obj is None:
            return obj
        if isinstance(obj, dict):
            return {k: sanitize_obj(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [sanitize_obj(v) for v in obj]
        return sanitize_value(obj)

    # Wrap overall import loop with robust error handling to avoid uncaught exceptions -> 500
    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            case_data = dict(zip(headers, row))
            # Sanitize raw data to be JSON stored safely
            case_data = sanitize_obj(case_data)
            # Map XLSX columns to Case fields (customize as needed)
            # Store the raw row data to allow frontend mapping and backfill
            title = case_data.get('Title') or case_data.get('title') or case_data.get('case_id') or 'No Title'
            description = case_data.get('Description') or case_data.get('description') or ''
            # Enforce system default status 'Pending' for imported cases
            status = 'Pending'
            # Store uploader info in raw so frontend can display who uploaded the record when available
            if isinstance(case_data, dict):
                case_data = dict(case_data)
            if uploader_user and isinstance(case_data, dict):
                # Avoid overwriting a raw uploaded_by if already present
                case_data.setdefault('uploaded_by', uploader_user.name)
            # create an import row record for tracking
            import_row = ImportRow(job_id=job.id, row_number=row_idx, raw=case_data, status='pending')
            db.add(import_row)
            db.flush()
            # Deduplicate: if raw contains an external identifier (case_id/_id/_uuid/caseNumber) that matches an existing case, skip
            dup_key = None
            for alias_key in ('case_id', '_id', '_uuid', 'caseNumber'):
                if alias_key in case_data and case_data.get(alias_key):
                    dup_key = str(case_data.get(alias_key))
                    break
            if dup_key:
                existing_case = None
                # Try database JSON path matching (preferred for PG/JSONB); but guard against dialects
                try:
                    existing_case = db.query(Case).filter(
                        or_(
                            Case.raw['case_id'].astext == dup_key,
                            Case.raw['_id'].astext == dup_key,
                            Case.raw['_uuid'].astext == dup_key,
                            Case.raw['caseNumber'].astext == dup_key,
                        )
                    ).first()
                except Exception as json_err:
                    # Fall back to Python-level scan (works for SQLite and other dialects)
                    logging.warning('JSON path query failed, falling back to python scan for dup key %s: %s', dup_key, json_err)
                    try:
                        rows_all = db.query(Case).filter(Case.raw != None).all()
                        for r in rows_all:
                            raw = r.raw or {}
                            try:
                                v = raw.get('case_id') or raw.get('_id') or raw.get('_uuid') or raw.get('caseNumber')
                            except Exception:
                                v = None
                            if v and str(v) == dup_key:
                                existing_case = r
                                break
                    except Exception as py_err:
                        logging.exception('Failed to fallback-scan cases for duplicate key %s: %s', dup_key, py_err)
                if existing_case:
                    import_row.case_id = existing_case.id
                    import_row.status = 'skipped'
                    db.add(import_row)
                    db.commit()
                    continue
            case = Case(
                title=title,
                description=description,
                status=status,
                raw=case_data,
            )
            # Do not assign uploader by default; keep system-unassigned when imported
            try:
                db.add(case)
                db.commit()
                db.refresh(case)
                imported += 1
                created_ids.append(case.id)
                # update the import row
                import_row.case_id = case.id
                import_row.status = 'success'
                db.add(import_row)
                db.commit()
            except Exception as e:
                logging.exception('Failed to import row: %s', e)
                db.rollback()
                # Update import_row to failed (best-effort)
                try:
                    import_row.error = str(e)
                    import_row.status = 'failed'
                    db.add(import_row)
                    db.commit()
                except Exception:
                    db.rollback()
                # record the failure so the client can show and reattempt
                failed_rows.append({'row': row_idx, 'error': str(e)})
                # skip faulty row and continue importing
                continue
    # db.commit() already performed per-row
    except HTTPException:
        # re-raise HTTPExceptions produced deliberately above
        raise
    except Exception as e:
        # Any unexpected error should be logged with job id if available
        logging.exception('Unhandled exception while processing uploaded file (import job id=%s): %s', getattr(job, 'id', 'N/A'), e)
        # Attempt to set job status / rows to failed for transparency
        try:
            job.status = 'failed'
            db.add(job)
            db.commit()
        except Exception:
            db.rollback()
        raise HTTPException(status_code=500, detail=f'Import failed due to server error: {e}')
    logging.info('Import summary: imported=%s created=%s failed=%s', imported, len(created_ids), len(failed_rows))
    return {"imported": imported, "created_ids": created_ids, "failed_rows": failed_rows, 'job_id': job.id}


@app.get('/import/jobs')
def list_import_jobs(db: Session = Depends(get_db)):
    jobs = db.query(ImportJob).order_by(ImportJob.created_at.desc()).all()
    results = []
    for j in jobs:
        total = len(j.rows)
        success = len([r for r in j.rows if r.status == 'success'])
        failed = len([r for r in j.rows if r.status == 'failed'])
        results.append({'id': j.id, 'uploader_name': j.uploader_name, 'filename': j.filename, 'created_at': j.created_at.isoformat(), 'total_rows': total, 'success': success, 'failed': failed})
    return results


@app.get('/import/jobs/{job_id}')
def get_import_job(job_id: int, db: Session = Depends(get_db), user=Depends(optional_auth)):
    job = db.get(ImportJob, job_id)
    if not job:
        raise HTTPException(status_code=404, detail='Job not found')
    sensitive_keys = set(['id_card_nu', 'family_card_nu', 'passport_nu_001', 'passaport_nu_001'])
    rows = []
    for r in job.rows:
        raw_content = r.raw if isinstance(r.raw, dict) else r.raw
        if not is_admin_user(user) and isinstance(raw_content, dict):
            sanitized = dict(raw_content)
            for k in list(sanitized.keys()):
                if k in sensitive_keys:
                    sanitized.pop(k, None)
            raw_content = sanitized
        rows.append({'row_number': r.row_number, 'status': r.status, 'error': r.error, 'case_id': r.case_id, 'raw': raw_content})
    return {'id': job.id, 'uploader_name': job.uploader_name, 'filename': job.filename, 'created_at': job.created_at.isoformat(), 'rows': rows}


@app.post('/import/jobs/{job_id}/retry')
def retry_import_job(job_id: int, db: Session = Depends(get_db), user=Depends(require_auth)):
    job = db.get(ImportJob, job_id)
    if not job:
        raise HTTPException(status_code=404, detail='Job not found')
    retry_count = 0
    for row in job.rows:
        if row.status == 'failed' or row.status == 'skipped':
            try:
                row_data = row.raw or {}
                title = row_data.get('Title') or row_data.get('title') or row_data.get('case_id') or 'No Title'
                description = row_data.get('Description') or row_data.get('description') or ''
                # dedupe by common keys; avoid creating duplicate case
                dup_key = None
                for alias_key in ('case_id', '_id', '_uuid', 'caseNumber'):
                    if alias_key in row_data and row_data.get(alias_key):
                        dup_key = str(row_data.get(alias_key))
                        break
                if dup_key:
                    existing = db.query(Case).filter(
                        or_(
                            Case.raw['case_id'].astext == dup_key,
                            Case.raw['_id'].astext == dup_key,
                            Case.raw['_uuid'].astext == dup_key,
                            Case.raw['caseNumber'].astext == dup_key,
                        )
                    ).first()
                    if existing:
                        row.case_id = existing.id
                        row.status = 'skipped'
                        db.add(row)
                        db.commit()
                        continue
                new_case = Case(title=title, description=description, status='Pending', raw=row_data)
                db.add(new_case)
                db.commit()
                db.refresh(new_case)
                row.case_id = new_case.id
                row.status = 'success'
                db.add(row)
                db.commit()
                retry_count += 1
            except Exception as e:
                db.rollback()
                row.error = str(e)
                row.status = 'failed'
                db.add(row)
                db.commit()
                continue
    return {'job_id': job.id, 'retried': retry_count}


@app.on_event('startup')
def on_startup():
    # quick check for DB connectivity
    try:
        db = SessionLocal()
        db.execute('SELECT 1')
        db.close()
    except Exception:
        # Log here if desired, but keep startup non-fatal
        pass

    # Seed default admin if no users exist (use env vars to override)
    try:
        db = SessionLocal()
        users_count = db.query(User).count()
        if users_count == 0:
            admin_user = os.getenv('INITIAL_ADMIN_USERNAME', 'admin')
            admin_email = os.getenv('INITIAL_ADMIN_EMAIL', 'admin@example.com')
            admin_pass = os.getenv('INITIAL_ADMIN_PASSWORD', 'admin123')
            hashed = hash_password(admin_pass)
            new_admin = User(username=admin_user, email=admin_email, password_hash=hashed, role='admin', name='Administrator')
            db.add(new_admin)
            db.commit()
            db.refresh(new_admin)
            # Log to keep track; avoid exposing password in logs
            print(f"[startup] created default admin user '{admin_user}' (email: {admin_email})")
        db.close()
    except Exception as e:
        print('[startup] error seeding admin user:', e)

# Simple auth: issue token for n8n or UI
@app.post("/auth/token")
def issue_token(payload: dict):
    # payload can include {"sub": "n8n", "role": "system"}
    # Keep legacy behavior (simple token creation for debugging) - default expiry
    token = create_token(payload)
    return {"token": token}


@app.post("/auth/generate")
def generate_token(payload: dict, user=Depends(require_auth)):
    """Generate a token for a given payload. Requires admin authorization.
    Accepts optional 'exp_minutes' in payload to set token TTL. Returns the signed token and expiry (minutes).
    """
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Admin privileges required to generate tokens")
    sub = payload.get('sub', 'n8n')
    role = payload.get('role', 'system')
    exp_minutes = payload.get('exp_minutes', None)
    token_payload = {"sub": sub, "role": role}
    if 'user_id' in payload:
        token_payload['user_id'] = payload.get('user_id')
    token = create_token(token_payload, exp_minutes=exp_minutes)
    return {"token": token, "expires_in_minutes": JWT_EXP_MINUTES if exp_minutes is None else int(exp_minutes)}

# Registration endpoint (no auth required - first user setup)
@app.post("/auth/register")
def register(payload: dict, db: Session = Depends(get_db)):
    username = payload.get("username")
    email = payload.get("email")
    password = payload.get("password")
    role = payload.get("role", "user")
    name = payload.get("name", username)
    
    if not username or not email or not password:
        raise HTTPException(status_code=400, detail="username, email, and password required")
    
    # Check if user exists
    existing = db.query(User).filter((User.username == username) | (User.email == email)).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username or email already exists")
    
    # Validate password strength during registration
    if not validate_password_strength(password):
        raise HTTPException(status_code=400, detail='Password does not meet strength requirements (min 8 chars; mixed case; digits or symbols)')
    # Create user
    user = User(
        username=username,
        email=email,
        password_hash=hash_password(password),
        role=role,
        name=name,
        ability=payload.get("ability")
    )
    if 'must_change_password' in payload:
        user.must_change_password = bool(payload.get('must_change_password'))
    db.add(user)
    db.commit()
    db.refresh(user)
    
    # Return token
    token = create_token({"sub": user.username, "user_id": user.id, "role": user.role})
    return {"token": token, "user": {"id": user.id, "username": user.username, "email": user.email, "role": user.role, "must_change_password": user.must_change_password if hasattr(user, 'must_change_password') else False}}

# Login endpoint
@app.post("/auth/login")
def login(payload: dict, db: Session = Depends(get_db)):
    try:
        username = payload.get("username")
        password = payload.get("password")
        
        if not username or not password:
            raise HTTPException(status_code=400, detail="username and password required")
        
        user = db.query(User).filter(User.username == username).first()
        if not user or not user.password_hash or not verify_password(password, user.password_hash):
            # Don't reveal which part failed (user exists vs password), just return unauthorized
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        token = create_token({"sub": user.username, "user_id": user.id, "role": user.role})
        return {"token": token, "user": {"id": user.id, "username": user.username, "email": user.email, "role": user.role, "must_change_password": user.must_change_password if hasattr(user, 'must_change_password') else False}}
    except HTTPException:
        # Pass-through known HTTP exceptions
        raise
    except Exception as e:
        logging.exception("Unhandled error in login for user '%s': %s", payload.get('username'), e)
        raise HTTPException(status_code=500, detail="Internal server error")


# Maintenance schedule endpoints (simple in-memory store)
@app.get('/maintenance')
def get_maintenance():
    return MAINTENANCE_SCHEDULES


@app.get('/cases/by-uploader/{uploader}')
def get_cases_by_uploader(uploader: str, db: Session = Depends(get_db)):
    # Return cases where `raw.uploaded_by` matches the uploader name (convenience for debugging/import verification)
    try:
        # Using JSON path query for uploaded_by
        results = db.query(Case).filter(Case.raw['uploaded_by'].astext == uploader).all()
        return results
    except Exception as e:
        logging.exception('Failed to query cases by uploader: %s', e)
        raise HTTPException(status_code=500, detail='Query failed')


@app.post('/maintenance')
def create_maintenance(payload: dict, user=Depends(require_auth)):
    global _maintenance_seq
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail='Admin privileges required')
    start = payload.get('start')
    end = payload.get('end')
    message = payload.get('message') or 'Scheduled maintenance'
    if not start or not end:
        raise HTTPException(status_code=400, detail='start and end datetime required')
    try:
        start_dt = datetime.fromisoformat(start)
        end_dt = datetime.fromisoformat(end)
    except Exception:
        raise HTTPException(status_code=400, detail='Invalid datetime format - use isoformat')
    if start_dt >= end_dt:
        raise HTTPException(status_code=400, detail='start must be before end')
    entry = {
        'id': _maintenance_seq,
        'start': start_dt.isoformat(),
        'end': end_dt.isoformat(),
        'message': message,
        'created_by': user.get('user_id') if isinstance(user, dict) else None,
    }
    _maintenance_seq += 1
    MAINTENANCE_SCHEDULES.append(entry)
    return entry


@app.delete('/maintenance/{mid}')
def delete_maintenance(mid: int, user=Depends(require_auth)):
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail='Admin privileges required')
    global MAINTENANCE_SCHEDULES
    MAINTENANCE_SCHEDULES = [m for m in MAINTENANCE_SCHEDULES if m['id'] != mid]
    return {'deleted': mid}
